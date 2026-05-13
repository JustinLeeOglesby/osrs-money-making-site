"""
OSRS Margin Tracker
Pulls live GE prices from the wiki API and writes a categorized spreadsheet
of profit margins for various conversions (herblore, crafting, fletching, cooking).

To extend: add new Recipe entries to the RECIPES list. Each recipe is just
inputs -> outputs with a category and subcategory tag.
"""

import requests
from dataclasses import dataclass, field
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===== CONFIG =====

HEADERS = {"User-Agent": "osrs-margin-tracker - personal use - your_contact_here"}
BASE_URL = "https://prices.runescape.wiki/api/v1/osrs"
OUTPUT_FILE = "osrs_margins.xlsx"

# Margin strategy:
#   "instant"  = buy inputs at high (insta-buy), sell outputs at low (insta-sell)
#                Conservative, what you'd realize doing it RIGHT NOW.
#   "patient"  = buy inputs at low, sell outputs at high
#                Optimistic, assumes your buy and sell offers both fill.
MARGIN_STRATEGY = "instant"


# ===== RECIPE MODEL =====


@dataclass
class Ingredient:
    item_id: int
    name: str
    quantity: int = 1


@dataclass
class Recipe:
    name: str
    category: str
    subcategory: str
    inputs: list[Ingredient]
    outputs: list[Ingredient]
    notes: str = ""
    is_f2p: bool = False
    xp: float = 0.0
    # Non-GE gp cost per craft (NPC fees, slayer master charges, etc.).
    # Folded into input_cost during margin calculation.
    extra_cost: int = 0
    # Free-form skill requirement, e.g. "Smithing 15" or "Magic 78".
    # Empty string = no requirement (or unknown).
    level_req: str = ""


# One way to perform a conversion. Pair with add_method_variants to emit
# parallel Recipes that share a subcategory so the methods sit adjacent
# in the output and can be compared directly.
@dataclass
class MethodVariant:
    method: str  # e.g. "Tanner", "Tan Leather spell"
    inputs: list[Ingredient]
    extra_cost: int = 0
    xp: float = 0.0
    notes: str = ""
    is_f2p: bool = False
    level_req: str = ""


def add_method_variants(
    base_name: str,
    category: str,
    output: Ingredient,
    variants: list[MethodVariant],
) -> None:
    """Emit one Recipe per variant. All variants share `base_name` as their
    subcategory so they group together in the spreadsheet/UI."""
    for v in variants:
        RECIPES.append(
            Recipe(
                name=f"{base_name} ({v.method})",
                category=category,
                subcategory=base_name,
                inputs=v.inputs,
                outputs=[output],
                notes=v.notes,
                is_f2p=v.is_f2p,
                xp=v.xp,
                extra_cost=v.extra_cost,
                level_req=v.level_req,
            )
        )


# Two-direction recipes: combine parts -> whole AND split whole -> parts.
# Split is an arbitrage signal even when the item isn't physically separable.
@dataclass
class Combination:
    name: str
    parts: list[Ingredient]
    whole: Ingredient
    notes: str = ""
    xp: float = 0.0  # XP awarded by the in-game Combine action (Split is always 0)
    category: str = "Combination items"
    subcategory: str = ""  # defaults to `name` so each combo's two directions sit together
    splittable: bool = True  # False = skip the reverse "Split" recipe entirely
    level_req: str = ""


# ===== ITEM IDS =====
# Grouped here for readability. Extend freely.

# Herbs: grimy / clean
GRIMY = {
    "Guam": 199,
    "Marrentill": 201,
    "Tarromin": 203,
    "Harralander": 205,
    "Ranarr": 207,
    "Toadflax": 3049,
    "Irit": 209,
    "Avantoe": 211,
    "Kwuarm": 213,
    "Snapdragon": 3051,
    "Cadantine": 215,
    "Lantadyme": 2485,
    "Dwarf weed": 217,
    "Torstol": 219,
}
CLEAN = {
    "Guam": 249,
    "Marrentill": 251,
    "Tarromin": 253,
    "Harralander": 255,
    "Ranarr": 257,
    "Toadflax": 2998,
    "Irit": 259,
    "Avantoe": 261,
    "Kwuarm": 263,
    "Snapdragon": 3000,
    "Cadantine": 265,
    "Lantadyme": 2481,
    "Dwarf weed": 267,
    "Torstol": 269,
}
# Unfinished potions (clean herb + vial of water)
UNF = {
    "Guam": 91,
    "Marrentill": 93,
    "Tarromin": 95,
    "Harralander": 97,
    "Ranarr": 99,
    "Toadflax": 3002,
    "Irit": 101,
    "Avantoe": 103,
    "Kwuarm": 105,
    "Snapdragon": 3004,
    "Cadantine": 107,
    "Lantadyme": 2483,
    "Dwarf weed": 109,
    "Torstol": 111,
}

VIAL_OF_WATER = 227

# Gems: uncut / cut
UNCUT_GEM = {
    "Sapphire": 1623,
    "Emerald": 1621,
    "Ruby": 1619,
    "Diamond": 1617,
    "Dragonstone": 1631,
    "Onyx": 6571,
}
CUT_GEM = {
    "Sapphire": 1607,
    "Emerald": 1605,
    "Ruby": 1603,
    "Diamond": 1601,
    "Dragonstone": 1615,
    "Onyx": 6573,
}

# Hides: raw / tanned
RAW_HIDE = {
    "Green d'hide": 1753,
    "Blue d'hide": 1751,
    "Red d'hide": 1749,
    "Black d'hide": 1747,
}
LEATHER = {
    "Green d'hide": 1745,
    "Blue d'hide": 2505,
    "Red d'hide": 2507,
    "Black d'hide": 2509,
}

# Fletching: logs / unstrung bows / strung bows
LOGS = {
    "Normal": 1511,
    "Oak": 1521,
    "Willow": 1519,
    "Maple": 1517,
    "Yew": 1515,
    "Magic": 1513,
}
UNSTRUNG_LONGBOW = {
    "Normal": 50,
    "Oak": 56,
    "Willow": 60,
    "Maple": 64,
    "Yew": 68,
    "Magic": 72,
}
LONGBOW = {
    "Normal": 839,
    "Oak": 843,
    "Willow": 847,
    "Maple": 851,
    "Yew": 855,
    "Magic": 859,
}
BOW_STRING = 1777

# Cooking: raw / cooked
RAW_FISH = {
    "Shrimp": 317,
    "Trout": 335,
    "Salmon": 331,
    "Tuna": 359,
    "Lobster": 377,
    "Swordfish": 371,
    "Monkfish": 7944,
    "Shark": 383,
    "Anglerfish": 13439,
    "Karambwan": 3142,
}
COOKED_FISH = {
    "Shrimp": 315,
    "Trout": 333,
    "Salmon": 329,
    "Tuna": 361,
    "Lobster": 379,
    "Swordfish": 373,
    "Monkfish": 7946,
    "Shark": 385,
    "Anglerfish": 13441,
    "Karambwan": 3144,
}


# Smithing: ores / bars (F2P)
ORE = {
    "Copper": 436,
    "Tin": 438,
    "Iron": 440,
    "Silver": 442,
    "Gold": 444,
    "Mithril": 447,
    "Adamantite": 449,
    "Runite": 451,
}
COAL = 453
BAR = {
    "Bronze": 2349,
    "Iron": 2351,
    "Steel": 2353,
    "Silver": 2355,
    "Gold": 2357,
    "Mithril": 2359,
    "Adamantite": 2361,
    "Runite": 2363,
}

# Other F2P crafting items
COWHIDE = 1739
SOFT_LEATHER = 1741
FLAX = 1779
WOOL = 1737
BALL_OF_WOOL = 1759

# Runes (for spell-based methods)
AIR_RUNE = 556
FIRE_RUNE = 554
WATER_RUNE = 555
EARTH_RUNE = 557
MIND_RUNE = 558
COSMIC_RUNE = 564
NATURE_RUNE = 561
LAW_RUNE = 563
DEATH_RUNE = 560
BLOOD_RUNE = 565
SOUL_RUNE = 566
ASTRAL_RUNE = 9075

# F2P crafting / fletching consumables
THREAD = 1734           # 5 uses per spool — treated as 1 per craft for simplicity
FEATHER = 314
ARROW_SHAFT = 52
HEADLESS_ARROW = 53


# ===== RECIPE DEFINITIONS =====

RECIPES: list[Recipe] = []

# Herblore: clean herbs
HERB_CLEAN_XP = {
    "Guam": 2.5, "Marrentill": 3.8, "Tarromin": 5.0, "Harralander": 6.3,
    "Ranarr": 7.5, "Toadflax": 8.0, "Irit": 8.8, "Avantoe": 10.0,
    "Kwuarm": 11.3, "Snapdragon": 11.8, "Cadantine": 12.5, "Lantadyme": 13.1,
    "Dwarf weed": 13.8, "Torstol": 15.0,
}
HERB_LEVEL = {
    "Guam": 3, "Marrentill": 5, "Tarromin": 11, "Harralander": 20,
    "Ranarr": 25, "Toadflax": 30, "Irit": 40, "Avantoe": 48,
    "Kwuarm": 54, "Snapdragon": 59, "Cadantine": 65, "Lantadyme": 67,
    "Dwarf weed": 70, "Torstol": 75,
}
for herb in GRIMY:
    RECIPES.append(
        Recipe(
            name=f"Clean {herb}",
            category="Herblore",
            subcategory="Clean herbs",
            inputs=[Ingredient(GRIMY[herb], f"Grimy {herb.lower()}")],
            outputs=[Ingredient(CLEAN[herb], f"Clean {herb.lower()}")],
            xp=HERB_CLEAN_XP[herb],
            level_req=f"Herblore {HERB_LEVEL[herb]}",
        )
    )

# Herblore: unfinished potions (clean herb + vial of water)
for herb in CLEAN:
    if herb in UNF:
        RECIPES.append(
            Recipe(
                name=f"{herb} unf potion",
                category="Herblore",
                subcategory="Unfinished potions",
                inputs=[
                    Ingredient(CLEAN[herb], f"Clean {herb.lower()}"),
                    Ingredient(VIAL_OF_WATER, "Vial of water"),
                ],
                outputs=[Ingredient(UNF[herb], f"{herb} potion (unf)")],
                level_req=f"Herblore {HERB_LEVEL[herb]}",
            )
        )

# Herblore: full chain — buy grimy + vial, sell unf potion. Skips the
# intermediate clean-herb GE round-trip (one fewer spread + one fewer tax).
for herb in GRIMY:
    if herb in UNF:
        RECIPES.append(
            Recipe(
                name=f"{herb} grimy to unf",
                category="Herblore",
                subcategory="Grimy to unf potion (full chain)",
                inputs=[
                    Ingredient(GRIMY[herb], f"Grimy {herb.lower()}"),
                    Ingredient(VIAL_OF_WATER, "Vial of water"),
                ],
                outputs=[Ingredient(UNF[herb], f"{herb} potion (unf)")],
                notes="Skips intermediate clean-herb buy/sell",
                xp=HERB_CLEAN_XP[herb],
                level_req=f"Herblore {HERB_LEVEL[herb]}",
            )
        )

# Crafting: gem cutting
F2P_CUT_GEMS = {"Sapphire", "Emerald", "Ruby", "Diamond"}
GEM_CUT_XP = {
    "Sapphire": 50.0, "Emerald": 67.5, "Ruby": 85.0, "Diamond": 107.5,
    "Dragonstone": 137.5, "Onyx": 167.5,
}
GEM_CUT_LEVEL = {
    "Sapphire": 20, "Emerald": 27, "Ruby": 34, "Diamond": 43,
    "Dragonstone": 55, "Onyx": 67,
}
for gem in UNCUT_GEM:
    RECIPES.append(
        Recipe(
            name=f"Cut {gem}",
            category="Crafting",
            subcategory="Gem cutting",
            inputs=[Ingredient(UNCUT_GEM[gem], f"Uncut {gem.lower()}")],
            outputs=[Ingredient(CUT_GEM[gem], gem)],
            is_f2p=gem in F2P_CUT_GEMS,
            xp=GEM_CUT_XP[gem],
            level_req=f"Crafting {GEM_CUT_LEVEL[gem]}",
        )
    )

# Crafting: hide tanning — two methods per hide, batched at 5 hides/leather
# so the Lunar Tan Leather spell's per-cast rune cost (1 astral + 1 nature
# + 1 fire, tans 5) lines up cleanly against paying the NPC tanner.
#   (display name, raw_id, leather_id, tanner_fee_gp, tanner_is_f2p)
HIDE_TANNINGS = [
    ("Cowhide", COWHIDE, SOFT_LEATHER, 1, True),
    ("Green d'hide", RAW_HIDE["Green d'hide"], LEATHER["Green d'hide"], 20, False),
    ("Blue d'hide", RAW_HIDE["Blue d'hide"], LEATHER["Blue d'hide"], 50, False),
    ("Red d'hide", RAW_HIDE["Red d'hide"], LEATHER["Red d'hide"], 100, False),
    ("Black d'hide", RAW_HIDE["Black d'hide"], LEATHER["Black d'hide"], 200, False),
]
TAN_LEATHER_SPELL_XP = 81.0  # Magic XP per cast (5 hides)
for hide_name, raw_id, leather_id, fee, tanner_f2p in HIDE_TANNINGS:
    leather_name = "Leather" if hide_name == "Cowhide" else f"{hide_name} leather"
    add_method_variants(
        base_name=f"Tan {hide_name}",
        category="Crafting",
        output=Ingredient(leather_id, leather_name, 5),
        variants=[
            MethodVariant(
                method="Tanner NPC",
                inputs=[Ingredient(raw_id, hide_name, 5)],
                extra_cost=fee * 5,
                notes=f"{fee}gp/hide tanner fee",
                is_f2p=tanner_f2p,
            ),
            MethodVariant(
                method="Tan Leather spell",
                inputs=[
                    Ingredient(raw_id, hide_name, 5),
                    Ingredient(ASTRAL_RUNE, "Astral rune", 2),
                    Ingredient(NATURE_RUNE, "Nature rune", 1),
                ],
                xp=TAN_LEATHER_SPELL_XP,
                notes="Lunar spell; XP shown is Magic XP per cast",
                level_req="Magic 78",
            ),
        ],
    )

# Fletching: logs -> unstrung longbows -> strung longbows
LONGBOW_FLETCH_XP = {
    "Normal": 10.0, "Oak": 25.0, "Willow": 41.5,
    "Maple": 58.0, "Yew": 67.5, "Magic": 91.5,
}
LONGBOW_LEVEL = {
    "Normal": 10, "Oak": 25, "Willow": 35,
    "Maple": 55, "Yew": 70, "Magic": 85,
}
for wood in LOGS:
    RECIPES.append(
        Recipe(
            name=f"{wood} longbow (u)",
            category="Fletching",
            subcategory="Cut bows",
            inputs=[Ingredient(LOGS[wood], f"{wood} logs")],
            outputs=[Ingredient(UNSTRUNG_LONGBOW[wood], f"{wood} longbow (u)")],
            xp=LONGBOW_FLETCH_XP[wood],
            level_req=f"Fletching {LONGBOW_LEVEL[wood]}",
        )
    )
    RECIPES.append(
        Recipe(
            name=f"String {wood} longbow",
            category="Fletching",
            subcategory="String bows",
            inputs=[
                Ingredient(UNSTRUNG_LONGBOW[wood], f"{wood} longbow (u)"),
                Ingredient(BOW_STRING, "Bow string"),
            ],
            outputs=[Ingredient(LONGBOW[wood], f"{wood} longbow")],
            xp=LONGBOW_FLETCH_XP[wood],
            level_req=f"Fletching {LONGBOW_LEVEL[wood]}",
        )
    )

# Fletching: finished darts (feather + metal dart tip -> dart). Modeled as a
# batch of 10 (which is how many tips a single bar produces in smithing) so
# the per-craft profit lines up with one bar's worth of inputs.
# (display_metal, tip_id, dart_id, level, xp_per_dart, is_f2p)
DART_RECIPES = [
    ("Bronze",   819,   806,   10, 1.8,  True),
    ("Iron",     820,   807,   22, 3.8,  True),
    ("Steel",    821,   808,   37, 7.5,  True),
    ("Mithril",  822,   809,   52, 11.2, True),
    ("Adamant",  823,   810,   67, 15.0, False),
    ("Rune",     824,   811,   81, 18.8, False),
    ("Amethyst", 25853, 25849, 90, 21.0, False),
    ("Dragon",   11232, 11230, 95, 25.0, False),
]
for metal, tip_id, dart_id, lvl, xp_per, is_f2p in DART_RECIPES:
    RECIPES.append(
        Recipe(
            name=f"{metal} darts",
            category="Fletching",
            subcategory="Darts",
            inputs=[
                Ingredient(tip_id, f"{metal} dart tip", 10),
                Ingredient(FEATHER, "Feather", 10),
            ],
            outputs=[Ingredient(dart_id, f"{metal} dart", 10)],
            xp=xp_per * 10,
            level_req=f"Fletching {lvl}",
            is_f2p=is_f2p,
            notes="10 per craft — feathering dart tips",
        )
    )

# Fletching: finished bolts (feather + unfinished bolt -> bolt). Standard
# inventory action makes 10 at once. Pairs with the unfinished-bolt entries
# in the Smithing tab to complete the bar->bolt chain.
# (display_metal, unf_id, bolt_id, level, xp_per_bolt, is_f2p)
FINISHED_BOLTS = [
    ("Bronze",  9375, 877,  9,  0.5,  True),
    ("Iron",    9377, 9140, 39, 1.5,  True),
    ("Steel",   9378, 9141, 46, 3.5,  True),
    ("Mithril", 9379, 9142, 54, 5.0,  True),
    ("Adamant", 9380, 9143, 61, 7.0,  False),
    ("Rune",    9381, 9144, 69, 10.0, False),
]
for metal, unf_id, bolt_id, lvl, xp_per, is_f2p in FINISHED_BOLTS:
    RECIPES.append(
        Recipe(
            name=f"{metal} bolts",
            category="Fletching",
            subcategory="Bolts (finished)",
            inputs=[
                Ingredient(unf_id, f"{metal} bolts (unf)", 10),
                Ingredient(FEATHER, "Feather", 10),
            ],
            outputs=[Ingredient(bolt_id, f"{metal} bolts", 10)],
            xp=xp_per * 10,
            level_req=f"Fletching {lvl}",
            is_f2p=is_f2p,
            notes="10 per craft — feathering unfinished bolts",
        )
    )

# Fletching: Amethyst arrows (newer 2018+ content — chiseled amethyst tips).
# Headless arrow + amethyst arrowtip -> amethyst arrow. 15 per craft, lvl 82.
AMETHYST_ARROWTIP = 21350
AMETHYST_ARROW = 21326
RECIPES.append(
    Recipe(
        name="Amethyst arrows",
        category="Fletching",
        subcategory="Arrows",
        inputs=[
            Ingredient(HEADLESS_ARROW, "Headless arrow", 15),
            Ingredient(AMETHYST_ARROWTIP, "Amethyst arrowtips", 15),
        ],
        outputs=[Ingredient(AMETHYST_ARROW, "Amethyst arrow", 15)],
        xp=13.5 * 15,
        level_req="Fletching 82",
        notes="15 per craft — top-tier ranged ammo, requires amethyst",
    )
)

# Fletching: Dragon arrows — headless arrow + dragon arrowtip -> dragon arrow.
DRAGON_ARROWTIP = 11237
DRAGON_ARROW = 11212
RECIPES.append(
    Recipe(
        name="Dragon arrows",
        category="Fletching",
        subcategory="Arrows",
        inputs=[
            Ingredient(HEADLESS_ARROW, "Headless arrow", 15),
            Ingredient(DRAGON_ARROWTIP, "Dragon arrowtips", 15),
        ],
        outputs=[Ingredient(DRAGON_ARROW, "Dragon arrow", 15)],
        xp=15.0 * 15,
        level_req="Fletching 90",
        notes="15 per craft — highest-tier standard arrow",
    )
)

# Fletching: Amethyst broad bolts — broad bolts + amethyst bolt tips -> tipped.
AMETHYST_BOLT_TIPS = 21944
BROAD_BOLTS = 13280
AMETHYST_BROAD_BOLTS = 21316
RECIPES.append(
    Recipe(
        name="Amethyst broad bolts",
        category="Fletching",
        subcategory="Tipped bolts",
        inputs=[
            Ingredient(BROAD_BOLTS, "Broad bolts", 10),
            Ingredient(AMETHYST_BOLT_TIPS, "Amethyst bolt tips", 10),
        ],
        outputs=[Ingredient(AMETHYST_BROAD_BOLTS, "Amethyst broad bolts", 10)],
        xp=10.6 * 10,
        level_req="Fletching 76",
        notes="10 per craft — broad bolts unlocked via Slayer rewards",
    )
)

# Fletching: Amethyst javelins — javelin shaft + amethyst javelin heads -> javelin.
JAVELIN_SHAFT = 19592
AMETHYST_JAVELIN_HEADS = 21358
AMETHYST_JAVELIN = 21318
RECIPES.append(
    Recipe(
        name="Amethyst javelins",
        category="Fletching",
        subcategory="Javelins",
        inputs=[
            Ingredient(JAVELIN_SHAFT, "Javelin shaft", 15),
            Ingredient(AMETHYST_JAVELIN_HEADS, "Amethyst javelin heads", 15),
        ],
        outputs=[Ingredient(AMETHYST_JAVELIN, "Amethyst javelin", 15)],
        xp=13.5 * 15,
        level_req="Fletching 84",
        notes="15 per craft — strong throwable ranged weapon",
    )
)

# Crafting: chiseling Amethyst into ammunition tips/heads. Raw amethyst is
# mined (lvl 92 Mining) and chiseled into one of four tip types. Wiki yields:
#   bolt tips:    15 per amethyst, lvl 83, 60 XP
#   arrowtips:    15 per amethyst, lvl 85, 60 XP
#   javelin heads: 5 per amethyst, lvl 87, 60 XP
#   dart tips:     8 per amethyst, lvl 89, 60 XP
AMETHYST = 21347
AMETHYST_DART_TIP = 25853
AMETHYST_CHISEL = [
    ("bolt tips",      AMETHYST_BOLT_TIPS,     15, 83),
    ("arrowtips",      AMETHYST_ARROWTIP,      15, 85),
    ("javelin heads",  AMETHYST_JAVELIN_HEADS,  5, 87),
    ("dart tips",      AMETHYST_DART_TIP,       8, 89),
]
for tip_name, tip_id, yield_qty, lvl in AMETHYST_CHISEL:
    RECIPES.append(
        Recipe(
            name=f"Chisel amethyst → {tip_name}",
            category="Crafting",
            subcategory="Amethyst chiseling",
            inputs=[Ingredient(AMETHYST, "Amethyst")],
            outputs=[Ingredient(tip_id, f"Amethyst {tip_name}", yield_qty)],
            xp=60.0,
            level_req=f"Crafting {lvl}",
            notes=f"{yield_qty} {tip_name} per amethyst",
        )
    )

# Cooking: raw -> cooked fish
F2P_COOK_FISH = {"Shrimp", "Trout", "Salmon", "Tuna", "Lobster", "Swordfish", "Shark"}
COOK_FISH_XP = {
    "Shrimp": 30.0, "Trout": 70.0, "Salmon": 90.0, "Tuna": 100.0,
    "Lobster": 120.0, "Swordfish": 140.0, "Monkfish": 150.0,
    "Shark": 210.0, "Anglerfish": 230.0, "Karambwan": 190.0,
}
COOK_FISH_LEVEL = {
    "Shrimp": 1, "Trout": 15, "Salmon": 25, "Tuna": 30,
    "Lobster": 40, "Swordfish": 45, "Monkfish": 62,
    "Shark": 80, "Anglerfish": 84, "Karambwan": 30,
}
for fish in RAW_FISH:
    RECIPES.append(
        Recipe(
            name=f"Cook {fish}",
            category="Cooking",
            subcategory="Fish",
            inputs=[Ingredient(RAW_FISH[fish], f"Raw {fish.lower()}")],
            outputs=[Ingredient(COOKED_FISH[fish], fish)],
            notes="Assumes no burns (high cooking lvl)",
            is_f2p=fish in F2P_COOK_FISH,
            xp=COOK_FISH_XP[fish],
            level_req=f"Cooking {COOK_FISH_LEVEL[fish]}",
        )
    )

# Smithing: smelt bars (all F2P)
# (display name, BAR key, [(ore key, qty), ...], coal qty, notes, smithing xp, level)
BAR_SMELT = [
    ("Bronze bar", "Bronze", [("Copper", 1), ("Tin", 1)], 0, "", 6.2, 1),
    ("Iron bar", "Iron", [("Iron", 1)], 0, "50% success without ring of forging", 12.5, 15),
    ("Steel bar", "Steel", [("Iron", 1)], 2, "", 17.5, 30),
    ("Silver bar", "Silver", [("Silver", 1)], 0, "", 13.7, 20),
    ("Gold bar", "Gold", [("Gold", 1)], 0, "Goldsmith gauntlets recommended", 22.5, 40),
    ("Mithril bar", "Mithril", [("Mithril", 1)], 4, "", 30.0, 50),
    ("Adamantite bar", "Adamantite", [("Adamantite", 1)], 6, "", 37.5, 70),
    ("Runite bar", "Runite", [("Runite", 1)], 8, "", 50.0, 85),
]
for bar_name, bar_key, ores, coal_qty, note, smelt_xp, smelt_lvl in BAR_SMELT:
    smelt_inputs = [
        Ingredient(ORE[ore_key], f"{ore_key} ore", qty) for ore_key, qty in ores
    ]
    if coal_qty > 0:
        smelt_inputs.append(Ingredient(COAL, "Coal", coal_qty))
    RECIPES.append(
        Recipe(
            name=f"Smelt {bar_name.lower()}",
            category="Smithing",
            subcategory="Smelt bars",
            inputs=smelt_inputs,
            outputs=[Ingredient(BAR[bar_key], bar_name)],
            notes=note,
            is_f2p=True,
            xp=smelt_xp,
            level_req=f"Smithing {smelt_lvl}",
        )
    )

# Smithing: bars -> items.
# Cannonballs are the classic AFK Smithing money method (steel bar + ammo
# mould near a furnace). Bolts (unfinished), dart tips, and knives are the
# other common bar-to-item conversions; profitability swings with bar prices.
CANNONBALL = 2
RECIPES.append(
    Recipe(
        name="Cannonballs",
        category="Smithing",
        subcategory="Cannonballs",
        inputs=[Ingredient(BAR["Steel"], "Steel bar")],
        outputs=[Ingredient(CANNONBALL, "Cannonball", 4)],
        xp=25.5,
        level_req="Smithing 35",
        notes="AFK staple — needs ammo mould (not consumed) + Dwarf Cannon quest",
    )
)

# Bolts (unfinished): 1 bar -> 10 unfinished bolts. Need feathers + Fletching
# later to finish, but the bar->unf step is pure Smithing arbitrage.
# (metal_key, item_id, level, xp_per_bar)
BOLTS_UNF = [
    ("Bronze",     9375,  9, 12.5),
    ("Iron",       9377, 39, 25.0),
    ("Steel",      9378, 53, 37.5),
    ("Mithril",    9379, 63, 50.0),
    ("Adamantite", 9380, 73, 62.5),
    ("Runite",     9381, 88, 75.0),
]
for metal, item_id, lvl, sxp in BOLTS_UNF:
    label = "Adamant" if metal == "Adamantite" else ("Rune" if metal == "Runite" else metal)
    RECIPES.append(
        Recipe(
            name=f"{label} bolts (unf)",
            category="Smithing",
            subcategory="Bolts (unfinished)",
            inputs=[Ingredient(BAR[metal], f"{metal} bar")],
            outputs=[Ingredient(item_id, f"{label} bolts (unf)", 10)],
            xp=sxp,
            level_req=f"Smithing {lvl}",
            is_f2p=(metal in {"Bronze", "Iron", "Steel", "Mithril"}),
        )
    )

# Dart tips: 1 bar -> 10 tips.
DART_TIPS = [
    ("Bronze",      819,  4, 1.0),
    ("Iron",        820, 22, 2.5),
    ("Steel",       821, 37, 3.75),
    ("Mithril",     822, 52, 5.0),
    ("Adamantite",  823, 67, 6.25),
    ("Runite",      824, 81, 7.5),
]
for metal, item_id, lvl, sxp in DART_TIPS:
    label = "Adamant" if metal == "Adamantite" else ("Rune" if metal == "Runite" else metal)
    RECIPES.append(
        Recipe(
            name=f"{label} dart tip",
            category="Smithing",
            subcategory="Dart tips",
            inputs=[Ingredient(BAR[metal], f"{metal} bar")],
            outputs=[Ingredient(item_id, f"{label} dart tip", 10)],
            xp=sxp,
            level_req=f"Smithing {lvl}",
            is_f2p=(metal in {"Bronze", "Iron", "Steel", "Mithril"}),
        )
    )

# Knives: 1 bar -> 5 knives.
KNIVES = [
    ("Bronze",     864,  4, 12.5),
    ("Iron",       863, 24, 25.0),
    ("Steel",      865, 39, 37.5),
    ("Mithril",    866, 59, 50.0),
    ("Adamantite", 867, 79, 62.5),
    ("Runite",     868, 94, 75.0),
]
for metal, item_id, lvl, sxp in KNIVES:
    label = "Adamant" if metal == "Adamantite" else ("Rune" if metal == "Runite" else metal)
    RECIPES.append(
        Recipe(
            name=f"{label} knife",
            category="Smithing",
            subcategory="Knives",
            inputs=[Ingredient(BAR[metal], f"{metal} bar")],
            outputs=[Ingredient(item_id, f"{label} knife", 5)],
            xp=sxp,
            level_req=f"Smithing {lvl}",
            is_f2p=(metal in {"Bronze", "Iron", "Steel", "Mithril"}),
        )
    )

# Crafting: spinning (F2P)
RECIPES.append(
    Recipe(
        name="Spin flax",
        category="Crafting",
        subcategory="Spinning",
        inputs=[Ingredient(FLAX, "Flax")],
        outputs=[Ingredient(BOW_STRING, "Bow string")],
        is_f2p=True,
        xp=15.0,
        level_req="Crafting 10",
    )
)
RECIPES.append(
    Recipe(
        name="Spin wool",
        category="Crafting",
        subcategory="Spinning",
        inputs=[Ingredient(WOOL, "Wool")],
        outputs=[Ingredient(BALL_OF_WOOL, "Ball of wool")],
        is_f2p=True,
        xp=2.5,
        level_req="Crafting 1",
    )
)

# Magic: Bones to Bananas (F2P, lvl 15) and Bones to Peaches (P2P, lvl 60).
# Each cast converts every bone in the inventory. Modeled per-cast assuming
# a full inv of 26 bones (2 slots for runes if not using staff/pouch). The
# spell ignores the actual rune cost if you carry a staff that provides the
# elemental runes — pricing here uses the strict rune cost.
BONES = 526
BANANA = 1963
PEACH = 6883
RECIPES.append(
    Recipe(
        name="Bones to Bananas (full inv)",
        category="Magic",
        subcategory="Bones to fruit",
        inputs=[
            Ingredient(BONES, "Bones", 26),
            Ingredient(EARTH_RUNE, "Earth rune", 2),
            Ingredient(WATER_RUNE, "Water rune", 2),
            Ingredient(NATURE_RUNE, "Nature rune", 1),
        ],
        outputs=[Ingredient(BANANA, "Banana", 26)],
        xp=15.0,
        level_req="Magic 15",
        is_f2p=True,
        notes="One cast converts all bones in inv; assumes 26-bone inventory",
    )
)
RECIPES.append(
    Recipe(
        name="Bones to Peaches (full inv)",
        category="Magic",
        subcategory="Bones to fruit",
        inputs=[
            Ingredient(BONES, "Bones", 26),
            Ingredient(EARTH_RUNE, "Earth rune", 4),
            Ingredient(WATER_RUNE, "Water rune", 4),
            Ingredient(NATURE_RUNE, "Nature rune", 2),
        ],
        outputs=[Ingredient(PEACH, "Peach", 26)],
        xp=35.5,
        level_req="Magic 60",
        notes="One cast converts all bones in inv; assumes 26-bone inventory",
    )
)

# Magic: Enchant Jewelry (Lvl 1-6 enchant spells).
# Every enchant cast also burns 1 cosmic rune on top of the level-specific
# runes. Inputs are the *finished* jewelry piece (cut gem already set in
# ring/necklace/bracelet; strung amulet for amulets), output is the enchanted
# variant. Recipes for pieces with no useful enchant target (e.g. ruby
# necklace, which only enchants into a digsite-pendant ingredient) are
# omitted. Bracelets are skipped here — several enchanted bracelet IDs are
# ambiguous; happy to add on request.
#
# (gem, magic_lvl, magic_xp, [(rune_id, qty), ...])
ENCHANT_SPELLS = {
    "Sapphire":    (7,  17.5, [(WATER_RUNE, 1)]),
    "Emerald":     (27, 37.0, [(AIR_RUNE, 3)]),
    "Ruby":        (49, 59.0, [(FIRE_RUNE, 5)]),
    "Diamond":     (57, 67.0, [(EARTH_RUNE, 10)]),
    "Dragonstone": (68, 78.0, [(EARTH_RUNE, 15), (WATER_RUNE, 15)]),
    "Onyx":        (87, 97.0, [(FIRE_RUNE, 20), (EARTH_RUNE, 20)]),
}
RUNE_NAMES = {
    AIR_RUNE: "Air rune", WATER_RUNE: "Water rune", EARTH_RUNE: "Earth rune",
    FIRE_RUNE: "Fire rune", COSMIC_RUNE: "Cosmic rune",
}

# (gem, piece_type, input_id, input_name, enchanted_id, enchanted_name)
ENCHANT_TARGETS = [
    # Rings -----------------------------------------------------------------
    ("Sapphire",    "Ring", 1637, "Sapphire ring",    2550, "Ring of recoil"),
    ("Emerald",     "Ring", 1639, "Emerald ring",     2552, "Ring of duelling(8)"),
    ("Ruby",        "Ring", 1641, "Ruby ring",        2568, "Ring of forging"),
    ("Diamond",     "Ring", 1643, "Diamond ring",     2570, "Ring of life"),
    ("Dragonstone", "Ring", 1645, "Dragonstone ring", 2572, "Ring of wealth"),
    ("Onyx",        "Ring", 6575, "Onyx ring",        6583, "Ring of stone"),
    # Amulets (strung) ------------------------------------------------------
    ("Sapphire",    "Amulet", 1694, "Sapphire amulet",    1727, "Amulet of magic"),
    ("Emerald",     "Amulet", 1696, "Emerald amulet",     1729, "Amulet of defence"),
    ("Ruby",        "Amulet", 1698, "Ruby amulet",        1725, "Amulet of strength"),
    ("Diamond",     "Amulet", 1700, "Diamond amulet",     1731, "Amulet of power"),
    ("Dragonstone", "Amulet", 1702, "Dragonstone amulet", 1712, "Amulet of glory"),
    ("Onyx",        "Amulet", 6581, "Onyx amulet",        6585, "Amulet of fury"),
    # Necklaces -------------------------------------------------------------
    ("Sapphire",    "Necklace", 1656, "Sapphire necklace",    3853,  "Games necklace(8)"),
    ("Emerald",     "Necklace", 1658, "Emerald necklace",     5521,  "Binding necklace"),
    ("Diamond",     "Necklace", 1662, "Diamond necklace",     11090, "Phoenix necklace"),
    ("Dragonstone", "Necklace", 1664, "Dragonstone necklace", 11105, "Skills necklace"),
    ("Onyx",        "Necklace", 6577, "Onyx necklace",        11128, "Berserker necklace"),
]
for gem, piece, in_id, in_name, out_id, out_name in ENCHANT_TARGETS:
    lvl, mxp, level_runes = ENCHANT_SPELLS[gem]
    inputs = [Ingredient(in_id, in_name), Ingredient(COSMIC_RUNE, "Cosmic rune")]
    for rune_id, qty in level_runes:
        inputs.append(Ingredient(rune_id, RUNE_NAMES[rune_id], qty))
    RECIPES.append(
        Recipe(
            name=f"Enchant {in_name.lower()} → {out_name}",
            category="Magic",
            subcategory=f"Enchant {piece.lower()}s",
            inputs=inputs,
            outputs=[Ingredient(out_id, out_name)],
            xp=mxp,
            level_req=f"Magic {lvl}",
            notes=f"Level-{list(ENCHANT_SPELLS).index(gem) + 1} Enchant spell",
        )
    )

# Crafting: leather items (F2P bank-standing — needle + thread + leather).
# Needle isn't consumed; thread is (~1 per item, simplified). Pure click-spam.
# (item_name, item_id, level, xp, leather_qty)
LEATHER_ITEMS = [
    ("Leather gloves",     1059,  1, 13.8, 1),
    ("Leather boots",      1061,  7, 16.3, 1),
    ("Leather cowl",       1167,  9, 18.5, 1),
    ("Leather vambraces",  1063, 11, 22.0, 1),
    ("Leather body",       1129, 14, 25.0, 1),
    ("Leather chaps",      1095, 18, 27.0, 1),
    ("Coif",               1169, 38, 37.0, 1),
]
for item_name, item_id, lvl, cxp, leather_qty in LEATHER_ITEMS:
    RECIPES.append(
        Recipe(
            name=item_name,
            category="Crafting",
            subcategory="Leather items",
            inputs=[
                Ingredient(SOFT_LEATHER, "Leather", leather_qty),
                Ingredient(THREAD, "Thread"),
            ],
            outputs=[Ingredient(item_id, item_name)],
            xp=cxp,
            level_req=f"Crafting {lvl}",
            is_f2p=True,
            notes="Bank-standing; needle is reusable",
        )
    )

# Fletching: headless arrows (F2P, lvl 1, very AFK bank-standing).
# Wiki: 15 made per inventory action; 1 xp each. Modeled per-batch of 15.
RECIPES.append(
    Recipe(
        name="Headless arrows",
        category="Fletching",
        subcategory="Arrows",
        inputs=[
            Ingredient(ARROW_SHAFT, "Arrow shaft", 15),
            Ingredient(FEATHER, "Feather", 15),
        ],
        outputs=[Ingredient(HEADLESS_ARROW, "Headless arrow", 15)],
        xp=15.0,
        level_req="Fletching 1",
        notes="15 per click — pure bank standing",
    )
)

# Fletching: arrow shafts from logs (F2P, lvl 1+; yield scales with log tier).
# Yields per the wiki: Normal=15, Oak=30, Willow=45, Maple=60, Yew=75, Magic=90
ARROW_SHAFT_DATA = [
    ("Normal", 1,  5.0,  15),
    ("Oak",    15, 10.0, 30),
    ("Willow", 30, 22.5, 45),
    ("Maple",  45, 36.0, 60),
    ("Yew",    60, 67.5, 75),
    ("Magic",  75, 91.5, 90),
]
for wood, lvl, fxp, yield_qty in ARROW_SHAFT_DATA:
    RECIPES.append(
        Recipe(
            name=f"{wood} arrow shafts",
            category="Fletching",
            subcategory="Arrow shafts",
            inputs=[Ingredient(LOGS[wood], f"{wood} logs")],
            outputs=[Ingredient(ARROW_SHAFT, "Arrow shaft", yield_qty)],
            xp=fxp,
            level_req=f"Fletching {lvl}",
        )
    )

# Crafting: battlestaves (orb + plain battlestaff -> elemental battlestaff).
# Classic bank-standing method — buy orbs and staves, click through inventory.
BATTLESTAFF = 1391
# (element, orb_id, output_id, level, xp)
BATTLESTAVES = [
    ("Water", 571, 1395, 54, 100.0),
    ("Earth", 575, 1399, 58, 112.5),
    ("Fire",  569, 1393, 62, 125.0),
    ("Air",   573, 1397, 66, 137.5),
]
for element, orb_id, out_id, lvl, cxp in BATTLESTAVES:
    RECIPES.append(
        Recipe(
            name=f"{element} battlestaff",
            category="Crafting",
            subcategory="Battlestaves",
            inputs=[
                Ingredient(BATTLESTAFF, "Battlestaff"),
                Ingredient(orb_id, f"{element} orb"),
            ],
            outputs=[Ingredient(out_id, f"{element} battlestaff")],
            xp=cxp,
            level_req=f"Crafting {lvl}",
            notes="Bank-standing: combine orb with battlestaff",
        )
    )

# Herblore: potion decanting. Combine lower-dose potions into (4)-dose and
# get the leftover empty vials back. Pure GE arbitrage — opens up whenever
# the per-dose price of an (n)-dose trades at a discount vs (4). No skill
# requirement, no XP.
#
# Restricted to commonly-traded potions; obscure ones (battlemage, bastion,
# overload, etc.) often have wide spreads or stale prices that make the
# margin numbers noisy rather than informative.
EMPTY_VIAL = 229  # "Vial" — the empty glass vial returned after decanting

# (display_name, {dose: item_id})
DECANT_POTIONS = [
    ("Prayer potion",       {1: 143,   2: 141,   3: 139,   4: 2434}),
    ("Super combat potion", {1: 12701, 2: 12699, 3: 12697, 4: 12695}),
    ("Saradomin brew",      {1: 6691,  2: 6689,  3: 6687,  4: 6685}),
    ("Super restore",       {1: 3030,  2: 3028,  3: 3026,  4: 3024}),
    ("Stamina potion",      {1: 12631, 2: 12629, 3: 12627, 4: 12625}),
    ("Ranging potion",      {1: 173,   2: 171,   3: 169,   4: 2444}),
    ("Magic potion",        {1: 3046,  2: 3044,  3: 3042,  4: 3040}),
    ("Antifire potion",     {1: 2458,  2: 2456,  3: 2454,  4: 2452}),
    ("Super antifire",      {1: 21986, 2: 21983, 3: 21980, 4: 21978}),
    ("Super attack",        {1: 149,   2: 147,   3: 145,   4: 2436}),
    ("Super strength",      {1: 161,   2: 159,   3: 157,   4: 2440}),
    ("Super defence",       {1: 167,   2: 165,   3: 163,   4: 2442}),
    ("Super energy",        {1: 3022,  2: 3020,  3: 3018,  4: 3016}),
    ("Energy potion",       {1: 3014,  2: 3012,  3: 3010,  4: 3008}),
    ("Antidote++",          {1: 5958,  2: 5956,  3: 5954,  4: 5952}),
]

# (input_dose, input_qty, output_4dose_qty, vial_returned_qty)
# Doses must conserve: in_dose * in_qty = 4 * out_qty.
DECANT_TRADES = [
    (1, 4, 1, 3),  # 4× (1) → 1× (4) + 3 vials
    (2, 2, 1, 1),  # 2× (2) → 1× (4) + 1 vial
    (3, 4, 3, 1),  # 4× (3) → 3× (4) + 1 vial
]

for potion_name, doses in DECANT_POTIONS:
    for in_dose, in_qty, out_qty, vial_qty in DECANT_TRADES:
        RECIPES.append(
            Recipe(
                name=f"Decant {potion_name.lower()} ({in_dose})→(4)",
                category="Decanting",
                subcategory=potion_name,
                inputs=[
                    Ingredient(doses[in_dose], f"{potion_name} ({in_dose})", in_qty),
                ],
                outputs=[
                    Ingredient(doses[4], f"{potion_name} (4)", out_qty),
                    Ingredient(EMPTY_VIAL, "Vial", vial_qty),
                ],
                notes="No skill / XP — pure GE arbitrage",
            )
        )


# Crafting: jewelry making (gold bar + optional cut gem on a furnace).
# Plain gold pieces use no gem. Mould cost is ignored (one-time, not consumed).
# Per-piece data: {gem_key: (item_id, crafting_level, xp)}; gem_key=None is the
# plain-gold variant.
RING_DATA = {
    None: (1635, 5, 15.0),
    "Sapphire": (1637, 20, 40.0),
    "Emerald": (1639, 27, 55.0),
    "Ruby": (1641, 34, 70.0),
    "Diamond": (1643, 43, 85.0),
    "Dragonstone": (1645, 55, 100.0),
    "Onyx": (6575, 67, 115.0),
}
NECKLACE_DATA = {
    None: (1654, 6, 20.0),
    "Sapphire": (1656, 22, 55.0),
    "Emerald": (1658, 29, 60.0),
    "Ruby": (1660, 40, 75.0),
    "Diamond": (1662, 56, 90.0),
    "Dragonstone": (1664, 72, 105.0),
    "Onyx": (6577, 82, 120.0),
}
BRACELET_DATA = {
    None: (11069, 7, 25.0),
    "Sapphire": (11072, 23, 60.0),
    "Emerald": (11076, 30, 65.0),
    "Ruby": (11085, 42, 80.0),
    "Diamond": (11092, 58, 95.0),
    "Dragonstone": (11115, 74, 110.0),
    "Onyx": (11130, 84, 125.0),
}
JEWELRY_TYPES = [
    # (piece_name, data, type_is_f2p)
    ("Ring", RING_DATA, True),
    ("Necklace", NECKLACE_DATA, True),
    ("Bracelet", BRACELET_DATA, False),  # bracelets are P2P regardless of gem
]
F2P_JEWELRY_GEMS = {None, "Sapphire", "Emerald", "Ruby", "Diamond"}
GOLD_BAR_ID = BAR["Gold"]

for piece_name, data, type_f2p in JEWELRY_TYPES:
    for gem in [None, "Sapphire", "Emerald", "Ruby", "Diamond", "Dragonstone", "Onyx"]:
        item_id, lvl, jxp = data[gem]
        if gem is None:
            display = f"Gold {piece_name.lower()}"
            inputs = [Ingredient(GOLD_BAR_ID, "Gold bar")]
        else:
            display = f"{gem} {piece_name.lower()}"
            inputs = [
                Ingredient(GOLD_BAR_ID, "Gold bar"),
                Ingredient(CUT_GEM[gem], gem),
            ]
        is_f2p = type_f2p and gem in F2P_JEWELRY_GEMS
        RECIPES.append(
            Recipe(
                name=display,
                category="Crafting",
                subcategory=f"Jewelry — {piece_name.lower()}s",
                inputs=inputs,
                outputs=[Ingredient(item_id, display)],
                xp=jxp,
                is_f2p=is_f2p,
                level_req=f"Crafting {lvl}",
            )
        )


# Farming: pot tree/fruit-tree seeds into saplings.
# Process: use seed on plant pot (filled with dirt), then water with a
# watering can. The plant pot stays with the seedling/sapling and is
# returned (empty) only when you eventually plant the sapling in a tree
# patch — so for the sapling-production margin, the pot is consumed.
# Watering-can charge cost is ignored (effectively free at any water source).
PLANT_POT_FILLED = 5354  # "Plant pot" with dirt; the watered/seeded result item
TREE_SAPLINGS = [
    # (seed_name, seed_id, sapling_name, sapling_id, farming_level, xp)
    ("Acorn", 5312, "Oak sapling", 5370, 15, 14.0),
    ("Willow seed", 5313, "Willow sapling", 5371, 30, 25.0),
    ("Maple seed", 5314, "Maple sapling", 5372, 45, 45.0),
    ("Yew seed", 5315, "Yew sapling", 5373, 60, 81.0),
    ("Magic seed", 5316, "Magic sapling", 5374, 75, 145.5),
]
FRUIT_TREE_SAPLINGS = [
    ("Apple tree seed", 5283, "Apple sapling", 5496, 27, 22.0),
    ("Banana tree seed", 5284, "Banana sapling", 5497, 33, 28.0),
    ("Orange tree seed", 5285, "Orange sapling", 5498, 39, 35.5),
    ("Curry tree seed", 5286, "Curry sapling", 5499, 42, 40.0),
    ("Pineapple seed", 5287, "Pineapple sapling", 5500, 51, 57.0),
    ("Papaya tree seed", 5288, "Papaya sapling", 5501, 57, 72.0),
    ("Palm tree seed", 5289, "Palm sapling", 5502, 68, 110.5),
    ("Dragonfruit tree seed", 22869, "Dragonfruit sapling", 22871, 81, 135.5),
]
for sap_group, sub in [
    (TREE_SAPLINGS, "Tree saplings"),
    (FRUIT_TREE_SAPLINGS, "Fruit tree saplings"),
]:
    for seed_name, seed_id, sap_name, sap_id, lvl, sap_xp in sap_group:
        RECIPES.append(
            Recipe(
                name=f"Pot {sap_name}",
                category="Farming",
                subcategory=sub,
                inputs=[
                    Ingredient(seed_id, seed_name),
                    Ingredient(PLANT_POT_FILLED, "Plant pot (filled)"),
                ],
                outputs=[Ingredient(sap_id, sap_name)],
                xp=sap_xp,
                level_req=f"Farming {lvl}",
                notes="Watering-can charge cost ignored",
            )
        )


# Combination items: each entry generates a Combine and a Split recipe so both
# directions show up adjacent on the spreadsheet.
GODSWORD_BLADE_ITEM = Ingredient(11798, "Godsword blade")

COMBINATIONS: list[Combination] = [
    Combination(
        name="Godsword blade",
        parts=[
            Ingredient(11818, "Godsword shard 1"),
            Ingredient(11820, "Godsword shard 2"),
            Ingredient(11822, "Godsword shard 3"),
        ],
        whole=GODSWORD_BLADE_ITEM,
        xp=100.0,
        level_req="Smithing 80",
    ),
    Combination(
        name="Armadyl godsword",
        parts=[GODSWORD_BLADE_ITEM, Ingredient(11810, "Armadyl hilt")],
        whole=Ingredient(11802, "Armadyl godsword"),
    ),
    Combination(
        name="Bandos godsword",
        parts=[GODSWORD_BLADE_ITEM, Ingredient(11812, "Bandos hilt")],
        whole=Ingredient(11804, "Bandos godsword"),
    ),
    Combination(
        name="Saradomin godsword",
        parts=[GODSWORD_BLADE_ITEM, Ingredient(11814, "Saradomin hilt")],
        whole=Ingredient(11806, "Saradomin godsword"),
    ),
    Combination(
        name="Zamorak godsword",
        parts=[GODSWORD_BLADE_ITEM, Ingredient(11816, "Zamorak hilt")],
        whole=Ingredient(11808, "Zamorak godsword"),
    ),
]

# Dragon bolts: feathered plain bolts, plus 10 gem-tipped variants.
# All under Fletching, all one-way (you can't decombine in-game).
DRAGON_BOLTS_PLAIN = Ingredient(21905, "Dragon bolts")
COMBINATIONS.append(
    Combination(
        name="Dragon bolts",
        parts=[
            Ingredient(21930, "Dragon bolts (unf)"),
            Ingredient(314, "Feather"),
        ],
        whole=DRAGON_BOLTS_PLAIN,
        notes="Add feathers to unfinished dragon bolts",
        xp=12.0,  # placeholder — verify against wiki
        category="Fletching",
        subcategory="Dragon bolts",
        splittable=False,
        level_req="Fletching 84",
    )
)
for gem_name, tip_id, tipped_id, tip_xp in [
    ("Opal", 45, 21955, 1.6),
    ("Jade", 9187, 21957, 1.8),
    ("Pearl", 46, 21959, 1.9),
    ("Topaz", 9188, 21961, 1.9),
    ("Sapphire", 9189, 21963, 2.5),
    ("Emerald", 9190, 21965, 2.65),
    ("Ruby", 9191, 21967, 2.8),
    ("Diamond", 9192, 21969, 2.7),
    ("Dragonstone", 9193, 21971, 2.8),
    ("Onyx", 9194, 21973, 3.0),
]:
    COMBINATIONS.append(
        Combination(
            name=f"{gem_name} dragon bolts",
            parts=[DRAGON_BOLTS_PLAIN, Ingredient(tip_id, f"{gem_name} bolt tips")],
            whole=Ingredient(tipped_id, f"{gem_name} dragon bolts"),
            notes="Unenchanted; (e) version requires Magic enchant spell + runes",
            xp=tip_xp,
            category="Fletching",
            subcategory="Dragon bolts",
            splittable=False,
        )
    )

# Wilderness wards (3 shards -> ward, no skill, 0 XP, one-way)
COMBINATIONS.extend([
    Combination(
        name="Malediction ward",
        parts=[
            Ingredient(11931, "Malediction shard 1"),
            Ingredient(11932, "Malediction shard 2"),
            Ingredient(11933, "Malediction shard 3"),
        ],
        whole=Ingredient(11924, "Malediction ward"),
        notes="Forge at the Volcanic Forge in the Wilderness",
        category="Combination items",
        subcategory="Wilderness wards",
        splittable=False,
    ),
    Combination(
        name="Odium ward",
        parts=[
            Ingredient(11928, "Odium shard 1"),
            Ingredient(11929, "Odium shard 2"),
            Ingredient(11930, "Odium shard 3"),
        ],
        whole=Ingredient(11926, "Odium ward"),
        notes="Forge at the Volcanic Forge in the Wilderness",
        category="Combination items",
        subcategory="Wilderness wards",
        splittable=False,
    ),
])

# Visage shields (Smithing 90, 2000 XP each, one-way)
COMBINATIONS.extend([
    Combination(
        name="Dragonfire shield",
        parts=[
            Ingredient(1540, "Anti-dragon shield"),
            Ingredient(11286, "Draconic visage"),
        ],
        whole=Ingredient(11284, "Dragonfire shield"),
        notes="Tradeable only when uncharged",
        xp=2000.0,
        category="Smithing",
        subcategory="Visage shields",
        splittable=False,
        level_req="Smithing 90",
    ),
    Combination(
        name="Dragonfire ward",
        parts=[
            Ingredient(1540, "Anti-dragon shield"),
            Ingredient(22006, "Skeletal visage"),
        ],
        whole=Ingredient(22003, "Dragonfire ward"),
        notes="Tradeable only when uncharged",
        xp=2000.0,
        category="Smithing",
        subcategory="Visage shields",
        splittable=False,
        level_req="Smithing 90",
    ),
])

# Blessed spirit shield (Prayer 85, 0 XP, one-way)
COMBINATIONS.append(
    Combination(
        name="Blessed spirit shield",
        parts=[
            Ingredient(12829, "Spirit shield"),
            Ingredient(12833, "Holy elixir"),
        ],
        whole=Ingredient(12831, "Blessed spirit shield"),
        category="Combination items",
        subcategory="Spirit shields",
        splittable=False,
        level_req="Prayer 85",
    )
)

# Trident of the swamp (Crafting 59, 0 XP, REVERSIBLE — chisel removes the fang)
COMBINATIONS.append(
    Combination(
        name="Trident of the swamp",
        parts=[
            Ingredient(11908, "Uncharged trident"),
            Ingredient(12932, "Magic fang"),
        ],
        whole=Ingredient(12900, "Uncharged toxic trident"),
        notes="Reversible with a chisel",
        category="Crafting",
        subcategory="Trident upgrade",
        splittable=True,
        level_req="Crafting 59",
    )
)


for combo in COMBINATIONS:
    sub = combo.subcategory or combo.name
    RECIPES.append(
        Recipe(
            name=f"Combine {combo.name}",
            category=combo.category,
            subcategory=sub,
            inputs=combo.parts,
            outputs=[combo.whole],
            notes=combo.notes,
            xp=combo.xp,
            level_req=combo.level_req,
        )
    )
    if combo.splittable:
        RECIPES.append(
            Recipe(
                name=f"Split {combo.name}",
                category=combo.category,
                subcategory=sub,
                inputs=[combo.whole],
                outputs=combo.parts,
                notes=combo.notes,
            )
        )

# Magic: enchant crossbow bolts (10 bolts per cast).
# Only the variants that were positive-margin at the time of writing are listed.
# Full table of spell costs / levels / XP for adding more later:
#   Opal        Magic  4  | 1 cosmic + 2 air                    |  9 XP/cast
#   Sapphire    Magic  7  | 1 cosmic + 1 water + 1 mind         | 17.5 XP
#   Jade        Magic 14  | 1 cosmic + 2 earth                  | 19 XP
#   Pearl       Magic 24  | 1 cosmic + 2 water                  | 29 XP
#   Emerald     Magic 27  | 1 cosmic + 3 air + 1 nature         | 37 XP
#   Topaz       Magic 29  | 1 cosmic + 2 fire                   | 33 XP
#   Ruby        Magic 49  | 1 cosmic + 5 fire + 1 blood         | 59 XP
#   Diamond     Magic 57  | 1 cosmic + 10 earth + 2 law         | 67 XP
#   Dragonstone Magic 68  | 1 cosmic + 15 earth + 1 soul        | 78 XP
#   Onyx        Magic 87  | 1 cosmic + 20 fire + 1 death        | 97 XP
RECIPES.append(
    Recipe(
        name="Enchant Opal dragon bolts",
        category="Magic",
        subcategory="Enchant crossbow bolts",
        inputs=[
            Ingredient(21955, "Opal dragon bolts", 10),
            Ingredient(COSMIC_RUNE, "Cosmic rune", 1),
            Ingredient(AIR_RUNE, "Air rune", 2),
        ],
        outputs=[Ingredient(21932, "Opal dragon bolts (e)", 10)],
        notes="Lucky Lightning effect; 10 bolts per cast",
        xp=9.0,
        level_req="Magic 4",
    )
)
RECIPES.append(
    Recipe(
        name="Enchant Ruby dragon bolts",
        category="Magic",
        subcategory="Enchant crossbow bolts",
        inputs=[
            Ingredient(21967, "Ruby dragon bolts", 10),
            Ingredient(COSMIC_RUNE, "Cosmic rune", 1),
            Ingredient(FIRE_RUNE, "Fire rune", 5),
            Ingredient(BLOOD_RUNE, "Blood rune", 1),
        ],
        outputs=[Ingredient(21944, "Ruby dragon bolts (e)", 10)],
        notes="Blood Forfeit effect; 10 bolts per cast",
        xp=59.0,
        level_req="Magic 49",
    )
)

# Magic + Fletching: full chain — buy plain dragon bolts + gem tips + runes,
# tip-then-enchant in one go, sell (e) bolts. Skips the GE round-trip on the
# tipped intermediate. xp field is Magic XP per cast; tipping Fletching XP
# is noted separately. Only includes currently profitable variants.
RECIPES.append(
    Recipe(
        name="Enchant Sapphire dragon bolts (full chain)",
        category="Magic",
        subcategory="Full chain (tip + enchant)",
        inputs=[
            Ingredient(21905, "Dragon bolts", 10),
            Ingredient(9189, "Sapphire bolt tips", 10),
            Ingredient(COSMIC_RUNE, "Cosmic rune", 1),
            Ingredient(WATER_RUNE, "Water rune", 1),
            Ingredient(MIND_RUNE, "Mind rune", 1),
        ],
        outputs=[Ingredient(21940, "Sapphire dragon bolts (e)", 10)],
        notes="Clear Mind effect; +25 Fletching XP for tipping; 10 bolts/cast",
        xp=17.5,
        level_req="Magic 7 / Fletching 56",
    )
)
RECIPES.append(
    Recipe(
        name="Enchant Opal dragon bolts (full chain)",
        category="Magic",
        subcategory="Full chain (tip + enchant)",
        inputs=[
            Ingredient(21905, "Dragon bolts", 10),
            Ingredient(45, "Opal bolt tips", 10),
            Ingredient(COSMIC_RUNE, "Cosmic rune", 1),
            Ingredient(AIR_RUNE, "Air rune", 2),
        ],
        outputs=[Ingredient(21932, "Opal dragon bolts (e)", 10)],
        notes="Lucky Lightning effect; +16 Fletching XP for tipping; 10 bolts/cast",
        xp=9.0,
        level_req="Magic 4 / Fletching 11",
    )
)
RECIPES.append(
    Recipe(
        name="Enchant Diamond dragon bolts (full chain)",
        category="Magic",
        subcategory="Full chain (tip + enchant)",
        inputs=[
            Ingredient(21905, "Dragon bolts", 10),
            Ingredient(9192, "Diamond bolt tips", 10),
            Ingredient(COSMIC_RUNE, "Cosmic rune", 1),
            Ingredient(EARTH_RUNE, "Earth rune", 10),
            Ingredient(LAW_RUNE, "Law rune", 2),
        ],
        outputs=[Ingredient(21946, "Diamond dragon bolts (e)", 10)],
        notes="Armour Piercing effect; +27 Fletching XP for tipping; 10 bolts/cast",
        xp=67.0,
        level_req="Magic 57 / Fletching 65",
    )
)
RECIPES.append(
    Recipe(
        name="Enchant Ruby dragon bolts (full chain)",
        category="Magic",
        subcategory="Full chain (tip + enchant)",
        inputs=[
            Ingredient(21905, "Dragon bolts", 10),
            Ingredient(9191, "Ruby bolt tips", 10),
            Ingredient(COSMIC_RUNE, "Cosmic rune", 1),
            Ingredient(FIRE_RUNE, "Fire rune", 5),
            Ingredient(BLOOD_RUNE, "Blood rune", 1),
        ],
        outputs=[Ingredient(21944, "Ruby dragon bolts (e)", 10)],
        notes="Blood Forfeit effect; +28 Fletching XP for tipping; 10 bolts/cast",
        xp=59.0,
        level_req="Magic 49 / Fletching 63",
    )
)


# ===== API + MATH =====


def fetch_latest_prices() -> dict:
    response = requests.get(f"{BASE_URL}/latest", headers=HEADERS, timeout=10)
    response.raise_for_status()
    return response.json()["data"]


def fetch_mapping() -> dict[int, dict]:
    """Returns {item_id: {'limit': X, 'name': Y, ...}}"""
    response = requests.get(f"{BASE_URL}/mapping", headers=HEADERS, timeout=10)
    response.raise_for_status()
    return {item["id"]: item for item in response.json()}


def calculate_ge_tax(sell_price: int) -> int:
    if sell_price < 50:
        return 0
    return min(sell_price // 100, 5_000_000)


def get_price(prices: dict, item_id: int, side: str) -> tuple[int | None, int | None]:
    """
    Returns (price, timestamp) for the given side ('high' or 'low').
    """
    item = prices.get(str(item_id))
    if not item:
        return None, None
    return item.get(side), item.get(f"{side}Time")


def calculate_recipe_margin(recipe: Recipe, prices: dict, mapping: dict) -> dict | None:
    """Compute margin for one recipe. Returns None if any price is missing."""
    if MARGIN_STRATEGY == "instant":
        input_side, output_side = "high", "low"
    else:
        input_side, output_side = "low", "high"

    input_lines = []
    input_cost = 0
    oldest_input_ts = None
    for ing in recipe.inputs:
        price, ts = get_price(prices, ing.item_id, input_side)
        if price is None:
            return None
        line_total = price * ing.quantity
        input_cost += line_total
        input_lines.append(
            {
                "item_id": ing.item_id,
                "name": ing.name,
                "qty": ing.quantity,
                "unit_price": price,
                "line_total": line_total,
            }
        )
        if ts is not None:
            oldest_input_ts = (
                ts if oldest_input_ts is None else min(oldest_input_ts, ts)
            )

    if recipe.extra_cost:
        input_cost += recipe.extra_cost
        input_lines.append(
            {
                "item_id": None,
                "name": "Coin cost (NPC fee)",
                "qty": 1,
                "unit_price": recipe.extra_cost,
                "line_total": recipe.extra_cost,
            }
        )

    output_lines = []
    output_revenue = 0
    total_tax = 0
    oldest_output_ts = None
    for out in recipe.outputs:
        price, ts = get_price(prices, out.item_id, output_side)
        if price is None:
            return None
        line_total = price * out.quantity
        line_tax = calculate_ge_tax(price) * out.quantity
        output_revenue += line_total
        total_tax += line_tax
        output_lines.append(
            {
                "item_id": out.item_id,
                "name": out.name,
                "qty": out.quantity,
                "unit_price": price,
                "line_total": line_total,
                "line_tax": line_tax,
            }
        )
        if ts is not None:
            oldest_output_ts = (
                ts if oldest_output_ts is None else min(oldest_output_ts, ts)
            )

    profit = output_revenue - input_cost - total_tax

    # GE buy limit of bottleneck input
    bottleneck_limit = None
    for ing in recipe.inputs:
        item_info = mapping.get(ing.item_id, {})
        limit = item_info.get("limit")
        if limit:
            # Convert to "max crafts per 4hr window" given quantity needed per craft
            crafts_possible = limit // ing.quantity
            if bottleneck_limit is None or crafts_possible < bottleneck_limit:
                bottleneck_limit = crafts_possible

    oldest_ts = min(filter(None, [oldest_input_ts, oldest_output_ts]), default=None)

    return {
        "input_cost": input_cost,
        "output_revenue": output_revenue,
        "tax": total_tax,
        "profit": profit,
        "buy_limit": bottleneck_limit,
        "max_4hr_profit": profit * bottleneck_limit if bottleneck_limit else None,
        "oldest_data_ts": oldest_ts,
        "input_lines": input_lines,
        "output_lines": output_lines,
    }


def age_minutes(ts: int | None) -> float | None:
    if ts is None:
        return None
    return (datetime.now(timezone.utc).timestamp() - ts) / 60


# ===== SPREADSHEET =====

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="305496")
SUBHEADER_FONT = Font(name="Arial", bold=True, size=11)
SUBHEADER_FILL = PatternFill("solid", start_color="D9E1F2")
DEFAULT_FONT = Font(name="Arial", size=10)
STALE_FILL = PatternFill("solid", start_color="FFC7CE")  # Pink for stale data
PROFIT_GOOD_FILL = PatternFill("solid", start_color="C6EFCE")  # Green for positive
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

COLUMNS = [
    ("Recipe", 28),
    ("Subcategory", 20),
    ("Level req", 14),
    ("Buy cost", 12),
    ("Sell revenue", 14),
    ("GE tax", 10),
    ("Profit / craft", 14),
    ("XP / craft", 11),
    ("GP / XP", 11),
    ("Buy limit (4hr)", 14),
    ("Max profit / 4hr", 16),
    ("Data age (min)", 14),
    ("Notes", 28),
]


def write_header_row(sheet, row: int):
    for col_idx, (label, width) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=row, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def write_recipe_row(sheet, row: int, recipe: Recipe, margin: dict):
    age = age_minutes(margin["oldest_data_ts"])
    is_stale = age is not None and age > 60
    gp_per_xp = margin["profit"] / recipe.xp if recipe.xp else None

    values = [
        recipe.name,
        recipe.subcategory,
        recipe.level_req or "—",
        margin["input_cost"],
        margin["output_revenue"],
        margin["tax"],
        margin["profit"],
        recipe.xp if recipe.xp else "—",
        round(gp_per_xp) if gp_per_xp is not None else "—",
        margin["buy_limit"] if margin["buy_limit"] is not None else "—",
        margin["max_4hr_profit"] if margin["max_4hr_profit"] is not None else "—",
        round(age, 1) if age is not None else "—",
        recipe.notes,
    ]

    for col_idx, value in enumerate(values, start=1):
        cell = sheet.cell(row=row, column=col_idx, value=value)
        cell.font = DEFAULT_FONT
        cell.border = THIN_BORDER
        # Currency / large-number columns
        if col_idx in (4, 5, 6, 7, 9, 11) and isinstance(value, (int, float)):
            cell.number_format = "#,##0;(#,##0);-"
        # XP column (one decimal)
        if col_idx == 8 and isinstance(value, (int, float)):
            cell.number_format = "#,##0.0"
        # Buy limit (positive integer)
        if col_idx == 10 and isinstance(value, int):
            cell.number_format = "#,##0"
        # Highlight profitable rows
        if col_idx == 7 and isinstance(value, (int, float)) and value > 0:
            cell.fill = PROFIT_GOOD_FILL
        # Highlight stale data
        if col_idx == 12 and is_stale:
            cell.fill = STALE_FILL


# Hidden by default; click the "+" outline button in the row gutter to expand
# and see how the recipe's profit breaks down per ingredient.
def write_recipe_detail_rows(sheet, start_row: int, margin: dict) -> int:
    row = start_row
    for line in margin["input_lines"]:
        sheet.cell(
            row=row,
            column=1,
            value=f"    Input: {line['qty']}x {line['name']} @ {line['unit_price']:,}",
        ).font = DEFAULT_FONT
        cost_cell = sheet.cell(row=row, column=4, value=line["line_total"])
        cost_cell.font = DEFAULT_FONT
        cost_cell.number_format = "#,##0"
        sheet.row_dimensions[row].outlineLevel = 1
        sheet.row_dimensions[row].hidden = True
        row += 1
    for line in margin["output_lines"]:
        sheet.cell(
            row=row,
            column=1,
            value=f"    Output: {line['qty']}x {line['name']} @ {line['unit_price']:,}",
        ).font = DEFAULT_FONT
        rev_cell = sheet.cell(row=row, column=5, value=line["line_total"])
        rev_cell.font = DEFAULT_FONT
        rev_cell.number_format = "#,##0"
        if line["line_tax"]:
            tax_cell = sheet.cell(row=row, column=6, value=line["line_tax"])
            tax_cell.font = DEFAULT_FONT
            tax_cell.number_format = "#,##0"
        sheet.row_dimensions[row].outlineLevel = 1
        sheet.row_dimensions[row].hidden = True
        row += 1
    return row


def build_category_sheet(wb: Workbook, category: str, recipes_with_margins: list):
    sheet = wb.create_sheet(category)
    sheet.sheet_properties.outlinePr.summaryBelow = False
    sheet.sheet_format.outlineLevelRow = 1
    # Sort by subcategory then by profit descending
    recipes_with_margins.sort(key=lambda x: (x[0].subcategory, -x[1]["profit"]))

    row = 1
    write_header_row(sheet, row)
    row += 1

    current_subcat = None
    for recipe, margin in recipes_with_margins:
        if recipe.subcategory != current_subcat:
            current_subcat = recipe.subcategory
            # Subcategory banner row
            for col_idx in range(1, len(COLUMNS) + 1):
                cell = sheet.cell(row=row, column=col_idx)
                cell.fill = SUBHEADER_FILL
                cell.border = THIN_BORDER
            sheet.cell(row=row, column=1, value=current_subcat).font = SUBHEADER_FONT
            row += 1
        write_recipe_row(sheet, row, recipe, margin)
        row += 1
        row = write_recipe_detail_rows(sheet, row, margin)

    sheet.freeze_panes = "A2"


def build_summary_sheet(wb: Workbook, all_results: list, generated_at: str):
    sheet = wb.create_sheet("Summary", 0)
    sheet.sheet_properties.outlinePr.summaryBelow = False
    sheet.sheet_format.outlineLevelRow = 1
    sheet["A1"] = "OSRS Margin Tracker"
    sheet["A1"].font = Font(name="Arial", bold=True, size=16)
    sheet["A2"] = f"Generated: {generated_at}"
    sheet["A2"].font = Font(name="Arial", italic=True, size=10)
    sheet["A3"] = (
        f"Strategy: {MARGIN_STRATEGY} (buy inputs {('high' if MARGIN_STRATEGY == 'instant' else 'low')}, sell outputs {('low' if MARGIN_STRATEGY == 'instant' else 'high')})"
    )
    sheet["A3"].font = Font(name="Arial", italic=True, size=10)

    # Top 25 most profitable per craft
    sheet["A5"] = "Top 25 by profit per craft"
    sheet["A5"].font = Font(name="Arial", bold=True, size=12)

    write_header_row(sheet, 6)
    sorted_all = sorted(all_results, key=lambda x: -x[1]["profit"])[:25]
    row = 7
    for recipe, margin in sorted_all:
        write_recipe_row(sheet, row, recipe, margin)
        # Override the subcategory column to show category for the summary view
        sheet.cell(row=row, column=2, value=f"{recipe.category} / {recipe.subcategory}")
        row += 1
        row = write_recipe_detail_rows(sheet, row, margin)

    sheet.freeze_panes = "A7"
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 32


def build_f2p_sheet(wb: Workbook, all_results: list):
    f2p_results = [(r, m) for r, m in all_results if r.is_f2p]
    if not f2p_results:
        return
    sheet = wb.create_sheet("F2P money makers", 1)
    sheet.sheet_properties.outlinePr.summaryBelow = False
    sheet.sheet_format.outlineLevelRow = 1
    f2p_results.sort(key=lambda x: (x[0].category, x[0].subcategory, -x[1]["profit"]))

    row = 1
    write_header_row(sheet, row)
    row += 1

    current_group = None
    for recipe, margin in f2p_results:
        group = (recipe.category, recipe.subcategory)
        if group != current_group:
            current_group = group
            for col_idx in range(1, len(COLUMNS) + 1):
                cell = sheet.cell(row=row, column=col_idx)
                cell.fill = SUBHEADER_FILL
                cell.border = THIN_BORDER
            sheet.cell(
                row=row, column=1, value=f"{recipe.category} - {recipe.subcategory}"
            ).font = SUBHEADER_FONT
            row += 1
        write_recipe_row(sheet, row, recipe, margin)
        row += 1
        row = write_recipe_detail_rows(sheet, row, margin)

    sheet.freeze_panes = "A2"


def main():
    print("Fetching prices and item mapping...")
    prices = fetch_latest_prices()
    mapping = fetch_mapping()

    print(f"Computing margins for {len(RECIPES)} recipes...")
    by_category: dict[str, list] = {}
    all_results = []
    skipped: list[str] = []
    for recipe in RECIPES:
        margin = calculate_recipe_margin(recipe, prices, mapping)
        if margin is None:
            unknown_ids = [
                f"{ing.name} (id={ing.item_id})"
                for ing in list(recipe.inputs) + list(recipe.outputs)
                if str(ing.item_id) not in prices
            ]
            reason = (
                "unknown IDs: " + ", ".join(unknown_ids)
                if unknown_ids
                else "no buy/sell price on the wiki right now"
            )
            skipped.append(f"{recipe.name} — {reason}")
            continue
        by_category.setdefault(recipe.category, []).append((recipe, margin))
        all_results.append((recipe, margin))

    if skipped:
        print(f"  Skipped {len(skipped)} recipes:")
        for s in skipped:
            print(f"    - {s}")

    print(f"Writing {OUTPUT_FILE}...")
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    build_summary_sheet(wb, all_results, generated_at)
    build_f2p_sheet(wb, all_results)
    for category in sorted(by_category):
        build_category_sheet(wb, category, by_category[category])

    wb.save(OUTPUT_FILE)
    print(
        f"Done. {len(all_results)} recipes priced across {len(by_category)} categories."
    )


if __name__ == "__main__":
    main()
